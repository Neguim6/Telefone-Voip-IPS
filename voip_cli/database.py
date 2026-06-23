import openpyxl
from pathlib import Path
from typing import Optional

COLUMNS = ["IP", "Setor", "Telefone", "Status", "Nome do Produto", "MAC", "Hardware", "Firmware", "Serial"]


class VoipDatabase:
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")

    def _load(self):
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(COLUMNS, row))
            if any(record.values()):
                data.append(record)
        return wb, ws, data

    def _save(self, wb):
        wb.save(self.filepath)

    def list_all(self) -> list[dict]:
        _, _, data = self._load()
        return data

    def search(self, **kwargs) -> list[dict]:
        _, _, data = self._load()
        results = []
        for record in data:
            for key, value in kwargs.items():
                if value and key in record:
                    if str(value).lower() in str(record[key] or "").lower():
                        results.append(record)
                        break
        return results

    def get_by_ip(self, ip: str) -> Optional[dict]:
        _, _, data = self._load()
        for record in data:
            if record["IP"] == ip:
                return record
        return None

    def add(self, record: dict) -> None:
        wb, ws, data = self._load()
        next_row = len(data) + 2
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=next_row, column=col_idx, value=record.get(col_name, ""))
        self._save(wb)

    def update(self, ip: str, updates: dict) -> bool:
        wb, ws, _ = self._load()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
            if row[0].value == ip:
                for col_idx, col_name in enumerate(COLUMNS, 1):
                    if col_name in updates:
                        ws.cell(row=row[0].row, column=col_idx, value=updates[col_name])
                self._save(wb)
                return True
        return False

    def delete(self, ip: str) -> bool:
        wb, ws, _ = self._load()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
            if row[0].value == ip:
                ws.delete_rows(row[0].row)
                self._save(wb)
                return True
        return False

    def stats(self) -> dict:
        data = self.list_all()
        total = len(data)
        online = sum(1 for r in data if str(r.get("Status", "")).upper() == "ONLINE")
        offline = sum(1 for r in data if str(r.get("Status", "")).upper() == "OFFLINE")
        modelos = {}
        setores = set()
        for r in data:
            m = r.get("Nome do Produto", "")
            if m:
                modelos[m] = modelos.get(m, 0) + 1
            s = r.get("Setor", "")
            if s:
                setores.add(s)
        return {
            "total": total,
            "online": online,
            "offline": offline,
            "modelos": modelos,
            "setores_unicos": len(setores),
        }
